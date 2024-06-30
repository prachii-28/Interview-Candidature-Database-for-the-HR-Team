import tkinter
from tkcalendar import Calendar
from tkinter import ttk
from tkinter import Toplevel, Button, END
from tkinter import messagebox
import os
import openpyxl
import pyodbc
import customtkinter
from CTkSpinbox import *

connection = pyodbc.connect('Driver={SQL Server};'
                            'Server=LAPTOP-HASKGEO3\SQLEXPRESS;'
                            'Database=SAMPLE_DB;'
                            'Trusted_Connection=yes;')
cursor = connection.cursor()


def show_reason(event):
    if result_combobox.get() == "Not Selected":
        reason_label.grid(row=8, column=1, padx=10, pady=5, sticky="news")
        reason_entry.grid(row=9, column=1, padx=10, pady=5, sticky="news")
    else:
        reason_label.grid_forget()
        reason_entry.grid_forget()


def clear_form():
    first_name_entry.delete(0, END)
    last_name_entry.delete(0, END)
    title_combobox.set('')
    email_id_entry.delete(0, END)
    phone_no_entry.delete(0, END)
    #age_spinbox.delete(0, END)
    age_spinbox.grid_forget()
    notice_period_combobox.set('')
    years_exp_entry.delete(0, END)
    last_interview_entry.delete(0, END)
    last_interview_entry.insert(0, "dd/mm/yyyy")
    skills_entry.delete(0, END)
    result_combobox.set('')
    reason_entry.delete(0, END)
    reason_label.grid_forget()
    reason_entry.grid_forget()


def submit():
    firstname = first_name_entry.get()
    lastname = last_name_entry.get()
    title = title_combobox.get()
    email_id = email_id_entry.get()
    phone_no = phone_no_entry.get()
    age = age_spinbox.get()
    notice_period = notice_period_combobox.get()
    experience = years_exp_entry.get()
    last_interview = last_interview_entry.get()
    skills = skills_entry.get()

    result = result_combobox.get()
    reason = reason_entry.get() if result == "Not Selected" else ""

    try:
        # Check for duplicacy
        check_sql = "SELECT Last_Date_Interviewed FROM Candidate_Data WHERE Phone_No = ?"
        cursor.execute(check_sql, phone_no)
        existing_candidate = cursor.fetchone()
        if existing_candidate:
           last_interview_date = existing_candidate[0]
           messagebox.showerror("Error", f"Candidate already exists! Last date interviewed: {last_interview_date}")
        else:

            insert_sql = '''
              INSERT INTO Candidate_Data (
              First_Name, Last_Name, Title, Email_id, Phone_No, Age,
              Notice_Period, Years_of_Experience, Last_Date_Interviewed, Skills, Result, Reason
              ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            '''
            cursor.execute(insert_sql, (
                firstname, lastname, title, email_id, phone_no, age,
                notice_period, experience, last_interview, skills, result, reason
            ))

            connection.commit()
            cursor.close()
            connection.close()

    # Optionally, show a messagebox on successful submission
            messagebox.showinfo("Success", "Candidate information submitted successfully!")

            clear_form()
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


    # main
    # filepath = "C:/Users/hp/Desktop/tkinter.xlsx"
    # workbook = openpyxl.load_workbook(filepath)
    # sheet = workbook.active

    # for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
    #     if phone_no in row:
    #         last_interview = sheet.cell(row=row[0], column=9).value
    #         messagebox.showerror("Error", f"Candidate already exists! Last date interviewed: {last_interview}")
    #         return

    # main
    # for row in range(2, sheet.max_row + 1):
    #     if phone_no == sheet.cell(row=row, column=5).value:
    #         last_interview = sheet.cell(row=row, column=9).value
    #         messagebox.showerror("Error", f"Candidate already exists! Last date interviewed: {last_interview}")
    #         return
    #
    # if not sheet['A1'].value:  # Check if the sheet is empty
    #     sheet.append(["First Name", "Last Name", "Title", "Email ID", "Phone No", "Age", "Notice Period", "Total Years of Experience", "Last Date Interviewed", "Skills", "Result", "Reason"])
    # sheet.append([firstname, lastname, title, email_id, phone_no, age, notice_period, experience, last_interview, skills, result, reason])
    # workbook.save(filepath)
customtkinter.set_appearance_mode('light')
customtkinter.set_default_color_theme("blue")
window = customtkinter.CTk()
window.title("Candidate Form")
window.geometry('600x650')
window.configure(bg='green')


def pick_date(event):
    global cal, date_window
    date_window = Toplevel()
    date_window.grab_set()
    date_window.title("select Date")
    date_window.geometry('250x220+590+370')
    cal = Calendar(date_window, selectmode="day", date_pattern="mm/dd/yyyy")
    cal.place(x=0, y=0)

    ok_bttn = Button(date_window, text="OK", command=grab_date)
    ok_bttn.place(x=80,y=190)


def grab_date():
    last_interview_entry.delete(0, END)
    last_interview_entry.insert(0, cal.get_date())
    date_window.destroy()


frame = customtkinter.CTkFrame(window, width=800, height=600, corner_radius=10)
frame.grid(row=1, column=1, sticky="nsew")
frame.pack(padx=20, pady=20)

# Saving Candidate Info
candidate_info_frame = tkinter.LabelFrame(frame, text="Candidate Information")
candidate_info_frame.grid(row=0 , column=0, padx=20, pady=10)

first_name_label = customtkinter.CTkLabel(candidate_info_frame, text="First Name")
first_name_label.grid(row=0, column=0, padx=20, pady=10)
first_name_entry = customtkinter.CTkEntry(candidate_info_frame, width=140, height=20)
first_name_entry.grid(row=1, column=0, padx=20, pady=10)

last_name_label = customtkinter.CTkLabel(candidate_info_frame, text="Last Name")
last_name_label.grid(row=0 , column=1, padx=20, pady=10)
last_name_entry = customtkinter.CTkEntry(candidate_info_frame, width=140, height=20)
last_name_entry.grid(row=1, column=1, padx=20, pady=10)

title_label = customtkinter.CTkLabel(candidate_info_frame, text="Title")
title_combobox = customtkinter.CTkComboBox(candidate_info_frame, values=["", "Ms. ", "Mr.", "Dr."], width=140, height=23)
title_label.grid(row=0, column=2, padx=20, pady=10)
title_combobox.grid(row=1, column=2, padx=20, pady=10)

email_id_label = customtkinter.CTkLabel(candidate_info_frame, text="Email ID")
email_id_label.grid(row=2, column=0, padx=20, pady=10)
email_id_entry = customtkinter.CTkEntry(candidate_info_frame, width=140, height=20)
email_id_entry.grid(row=3, column=0, padx=20, pady=10)

phone_no_label = customtkinter.CTkLabel(candidate_info_frame, text="Phone No")
phone_no_label.grid(row=2, column=1, padx=20, pady=10)
phone_no_entry = customtkinter.CTkEntry(candidate_info_frame, width=140, height=20)
phone_no_entry.grid(row=3, column=1, padx=20, pady=10)

age_label = customtkinter.CTkLabel(candidate_info_frame, text="Age")
age_spinbox = CTkSpinbox(candidate_info_frame, start_value=18, min_value=18, max_value=60, width=140, height=23)
age_label.grid(row=2, column=2, padx=20, pady=10)
age_spinbox.grid(row=3, column=2, padx=20, pady=10)

# saving other info
other_info_frame = tkinter.LabelFrame(frame)
other_info_frame.grid(row=1, column=0, sticky="news", padx=20, pady=20)

notice_period_label = customtkinter.CTkLabel(other_info_frame, text="Notice Period")
notice_period_label.grid(row=0, column=0, padx=20, pady=10)
notice_period_combobox = customtkinter.CTkComboBox(other_info_frame, values=["available", "15 days", "30 days", "60 days", "90 days"], width=140, height=23)
notice_period_combobox.set('')
notice_period_combobox.grid(row=1, column=0, padx=20, pady=10)

years_exp_label = customtkinter.CTkLabel(other_info_frame, text="Total Years of Experience")
years_exp_label.grid(row=0, column=1, padx=20, pady=10)
years_exp_entry = customtkinter.CTkEntry(other_info_frame, width=140, height=20)
years_exp_entry.grid(row=1,column=1, padx=20, pady=10)

last_interview_label = customtkinter.CTkLabel(other_info_frame, text="Last Date Interviewed")
last_interview_label.grid(row=0, column=2, padx=20, pady=10)
last_interview_entry = customtkinter.CTkEntry(other_info_frame, width=140, height=20)
last_interview_entry.grid(row=1,column=2, padx=20, pady=10)
last_interview_entry.insert(0, "dd/mm/yyyy")
last_interview_entry.bind("<1>", pick_date)

# skills labelframe
skills_frame = tkinter.LabelFrame(frame)
skills_frame.grid(row=2, column=0, sticky="news", padx=20, pady=20)
skills_frame.grid_rowconfigure(0, weight=1)
skills_frame.grid_rowconfigure(1, weight=1)
skills_frame.grid_columnconfigure(0, weight=1)
skills_frame.grid_columnconfigure(1, weight=1)
skills_frame.grid_columnconfigure(2, weight=1)

skills_label = customtkinter.CTkLabel(skills_frame, text="Enter your Skills")
skills_label.grid(row=0, column=1, padx=20, pady=10, sticky="ew")
skills_entry = customtkinter.CTkEntry(skills_frame, width=180, height=20)
skills_entry.grid(row=1, column=1, padx=20, pady=10, sticky="ew")


result_frame = tkinter.LabelFrame(frame)
result_frame.grid(row=3, column=0, sticky="news", padx=20, pady=20)

result_frame.grid_rowconfigure(0, weight=1)
result_frame.grid_rowconfigure(1, weight=1)
result_frame.grid_rowconfigure(2, weight=1)
result_frame.grid_rowconfigure(3, weight=1)
result_frame.grid_columnconfigure(0, weight=1)
result_frame.grid_columnconfigure(1, weight=1)
result_frame.grid_columnconfigure(2, weight=1)
result_label = customtkinter.CTkLabel(result_frame, text="Result")
result_label.grid(row=0, column=1, padx=20, pady=10)
result_combobox = customtkinter.CTkComboBox(result_frame, values=["Selected", "Not Selected"], width=140, height=23)
result_combobox.set('')
result_combobox.grid(row=1, column=1, padx=20, pady=10)
result_combobox.bind("<<ComboboxSelected>>", show_reason)

reason_label = customtkinter.CTkLabel(result_frame, text="Reason for Rejection:")
reason_label.grid(row=2, column=1, padx=20, pady=10)
reason_entry = customtkinter.CTkEntry(result_frame, width=140, height=20)
reason_entry.grid(row=3, column=1, padx=20, pady=10)


for widget in candidate_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

for widget in other_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

for widget in result_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# Button
button = customtkinter.CTkButton(frame, text="Submit",command=submit)
button.grid(row=9, column=0, padx=20, pady=10)

window.mainloop()
